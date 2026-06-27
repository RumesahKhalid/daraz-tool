from flask import Flask, request, jsonify, send_file
from flask_cors import CORS
import anthropic
import openpyxl
from openpyxl import load_workbook
import base64
import json
import io
import os

app = Flask(__name__)
CORS(app, origins="*", allow_headers=["Content-Type"], methods=["GET", "POST", "OPTIONS"])

client = anthropic.Anthropic(api_key=os.environ.get("ANTHROPIC_API_KEY"))

@app.route('/')
def home():
    return "Daraz Tool Backend Running!"

@app.route('/analyze', methods=['POST'])
def analyze():
    try:
        data = request.json
        image_base64 = data.get('image')
        
        message = client.messages.create(
            model="claude-sonnet-4-6",
            max_tokens=1000,
            system="""You are a Daraz Pakistan product listing expert. Analyze the product image and return ONLY valid JSON with these exact fields:
{
  "productName": "20-40 char SEO name with brand/type/feature",
  "brand": "brand name or Generic if not visible",
  "color": "one color from: Black, White, Red, Blue, Green, Yellow, Pink, Purple, Grey, Silver, Gold, Multicolor, Brown, Orange, Beige",
  "capacity": "e.g. 1L, 500ml, 2kg or empty string if not applicable",
  "highlights": "3-5 bullet points separated by newlines starting with bullet about key features",
  "description": "2-3 paragraph HTML description for Pakistani buyers"
}
Respond with ONLY the JSON object. No markdown, no explanation.""",
            messages=[{
                "role": "user",
                "content": [
                    {
                        "type": "image",
                        "source": {
                            "type": "base64",
                            "media_type": "image/jpeg",
                            "data": image_base64
                        }
                    },
                    {
                        "type": "text",
                        "text": "Analyze this product image and return the JSON listing details."
                    }
                ]
            }]
        )
        
        text = message.content[0].text.replace('```json', '').replace('```', '').strip()
        result = json.loads(text)
        return jsonify({"success": True, "data": result})
        
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500


@app.route('/generate-excel', methods=['POST'])
def generate_excel():
    try:
        data = request.json
        products = data.get('products', [])
        template_data = data.get('template')
        category_name = data.get('categoryName', 'Sheet1')

        template_bytes = base64.b64decode(template_data)
        wb = load_workbook(io.BytesIO(template_bytes))
        
        ws = None
        for sheet_name in wb.sheetnames:
            if sheet_name.lower() not in ['index', 'processresult', 'global_hide'] and 'hide' not in sheet_name.lower():
                ws = wb[sheet_name]
                break
        
        if ws is None:
            return jsonify({"success": False, "error": "Template sheet not found"}), 400

        header_row = None
        for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=10), 1):
            for cell in row:
                if cell.value and '*Product Name' in str(cell.value):
                    header_row = row_idx
                    break
            if header_row:
                break

        if not header_row:
            return jsonify({"success": False, "error": "Header row not found"}), 400

        col_map = {}
        for cell in ws[header_row]:
            if cell.value:
                col_map[str(cell.value).strip()] = cell.column

        start_row = header_row + 1
        for ws_row in ws.iter_rows(min_row=start_row, max_row=ws.max_row):
            for cell in ws_row:
                cell.value = None

        for i, product in enumerate(products):
            row_num = start_row + i
            
            def set_col(header, value):
                col = col_map.get(header)
                if col:
                    ws.cell(row=row_num, column=col, value=value)

            set_col('*Product Name(English)', product.get('productName', ''))
            set_col('*Brand', product.get('brand', 'Generic'))
            set_col('capacity', product.get('capacity', ''))
            set_col('*Highlights', product.get('highlights', ''))
            set_col('*Product Description', product.get('description', ''))
            set_col("What's in the box", '1 x ' + product.get('productName', 'Product'))
            set_col('*Package Weight (kg)', 0.02)
            set_col('*Package Length(cm) * Width(cm) * Height(cm)-Length (cm)', 8)
            set_col('*Package Length(cm) * Width(cm) * Height(cm)-Width (cm)', 8)
            set_col('*Package Length(cm) * Width(cm) * Height(cm)-Height (cm)', 8)
            set_col('Dangerous Goods', 'None')
            set_col('Standard', 'Yes')
            set_col('*Color family', product.get('color', 'Not Specified'))
            set_col('*Quantity', 100)
            set_col('*Price', product.get('price', 0))
            if product.get('specialPrice'):
                set_col('SpecialPrice', product.get('specialPrice'))
            set_col('Warranty', 'No Warranty')
            set_col('Warranty Type', 'No Warranty')

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'daraz_{category_name}.xlsx'
        )

    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500


if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    app.run(host='0.0.0.0', port=port)
